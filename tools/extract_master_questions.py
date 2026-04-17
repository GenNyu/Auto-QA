#!/usr/bin/env python3
import argparse
import json
import os
from typing import Optional, List, Any

import openpyxl


def find_header_row(ws, header: str, max_rows: int = 30, max_cols: int = 50) -> Optional[int]:
    target = header.strip().lower()
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == target:
                return r
    return None


def find_header_col(ws, header_row: int, header: str) -> Optional[int]:
    target = header.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().lower() == target:
            return c
    return None


def normalize_question(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def main():
    parser = argparse.ArgumentParser(description="Extract 'Question' column and add ids")
    parser.add_argument(
        "-i",
        "--input",
        default="input/source/Master Data.xlsx",
        help="Input .xlsx file",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="outputs/Master_Data_questions.json",
        help="Output .json file",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        default="Master_KB",
        help="Sheet name (default: Master_KB; falls back to active if not found)",
    )
    parser.add_argument(
        "--header",
        default="Question",
        help="Header name for the question column",
    )
    parser.add_argument(
        "--id-width",
        type=int,
        default=3,
        help="Zero-pad width for id (default: 3)",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input)
    if args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb.active

    header_row = find_header_row(ws, args.header)
    if header_row is None:
        raise ValueError(f"Header '{args.header}' not found in sheet '{ws.title}'")

    col = find_header_col(ws, header_row, args.header)
    if col is None:
        raise ValueError(f"Column for header '{args.header}' not found in sheet '{ws.title}'")

    questions: List[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        q = normalize_question(ws.cell(r, col).value)
        if q is None:
            continue
        questions.append(q)

    width = max(args.id_width, len(str(len(questions))))
    data = [
        {
            "id": str(i + 1).zfill(width),
            "question": q,
        }
        for i, q in enumerate(questions)
    ]

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Wrote {args.output} ({len(data)} items)")


if __name__ == "__main__":
    main()
