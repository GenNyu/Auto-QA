#!/usr/bin/env python3
import argparse
import os
import json
from typing import Optional, Dict, List

import openpyxl


def find_header_row(ws, target: str, max_rows: int = 25, max_cols: int = 25) -> Optional[int]:
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == target:
                return r
    return None


def build_merged_lookup(ws) -> List[openpyxl.worksheet.merge.MergedCellRange]:
    return list(ws.merged_cells.ranges)


def get_cell_with_merged(ws, merged_ranges, row: int, col: int):
    v = ws.cell(row, col).value
    if v is not None:
        return v
    for rng in merged_ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col).value
    return None


def extract_sheet(ws):
    header_row = find_header_row(ws, "PCI DSS Requirements")
    if header_row is None:
        raise ValueError(f"Header row not found in sheet '{ws.title}'")

    merged_ranges = build_merged_lookup(ws)
    max_col = ws.max_column

    top_headers: Dict[int, Optional[str]] = {}
    sub_headers: Dict[int, Optional[str]] = {}

    for c in range(1, max_col + 1):
        top_headers[c] = get_cell_with_merged(ws, merged_ranges, header_row, c)
        sub_headers[c] = get_cell_with_merged(ws, merged_ranges, header_row + 1, c)

    def find_col(top_name: str, sub_name: Optional[str] = None) -> Optional[int]:
        for c in range(1, max_col + 1):
            if top_headers[c] == top_name:
                if sub_name is None or sub_headers[c] == sub_name:
                    return c
        return None

    col_pci = find_col("PCI DSS Requirements")
    col_test = find_col("Testing Procedures")
    col_impl_customer = find_col("Implementation Details", "Customer Only")

    if col_pci is None or col_test is None or col_impl_customer is None:
        raise ValueError(
            f"Required columns not found in sheet '{ws.title}'. "
            f"Found PCI={col_pci}, Testing={col_test}, Impl.Customer={col_impl_customer}"
        )

    data_rows = []
    for r in range(header_row + 2, ws.max_row + 1):
        row_vals = [
            # Inherit merged PCI DSS Requirements for sub-rows
            get_cell_with_merged(ws, merged_ranges, r, col_pci),
            ws.cell(r, col_test).value,
            # Handle merged cells so sub-rows inherit the same Implementation Details
            get_cell_with_merged(ws, merged_ranges, r, col_impl_customer),
        ]
        if all(v is None for v in row_vals):
            continue
        data_rows.append(row_vals)

    return data_rows


def main():
    parser = argparse.ArgumentParser(description="Filter PCI DSS columns from Excel workbook")
    parser.add_argument(
        "-i",
        "--input",
        default="input/source/PCI_DSS_Controls_Responsibility_Matrix_UrBox_Visa_FINAL.xlsx",
        help="Input .xlsx file",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="outputs/PCI_DSS_filtered.xlsx",
        help="Output .xlsx file",
    )
    parser.add_argument(
        "--format",
        choices=["xlsx", "json"],
        default="xlsx",
        help="Output format",
    )
    args = parser.parse_args()

    in_path = args.input
    out_path = args.output
    out_format = args.format

    wb = openpyxl.load_workbook(in_path)
    if out_format == "xlsx":
        out_wb = openpyxl.Workbook()
        # remove default sheet
        out_wb.remove(out_wb.active)
    else:
        all_rows = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        data_rows = extract_sheet(ws)

        if out_format == "xlsx":
            out_ws = out_wb.create_sheet(title=sheet[:31])
            out_ws.append([
                "PCI DSS Requirements",
                "Testing Procedures",
                "Implementation Details - Customer Only",
            ])
            for row in data_rows:
                out_ws.append(row)
        else:
            for row in data_rows:
                all_rows.append({
                    "req": sheet,
                    "pci_dss_requirements": row[0],
                    "testing_procedures": row[1]
                })

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    if out_format == "xlsx":
        out_wb.save(out_path)
    else:
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, ensure_ascii=False, indent=2)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
