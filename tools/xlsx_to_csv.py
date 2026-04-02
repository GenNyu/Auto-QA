#!/usr/bin/env python3
"""Convert an XLSX file to CSV.

Usage examples:
  python xlsx_to_csv.py "Master Data.xlsx"
  python xlsx_to_csv.py "Master Data.xlsx" --sheet "Sheet1" --out "Master Data.csv"
  python xlsx_to_csv.py "Master Data.xlsx" --out-dir ./csv
"""

import argparse
import csv
import os
import sys

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - runtime dependency
    print(
        "Missing dependency: openpyxl. Install with: pip install openpyxl",
        file=sys.stderr,
    )
    raise


def _safe_filename(name: str) -> str:
    keep = []
    for ch in name:
        if ch.isalnum() or ch in (" ", "-", "_", "."):
            keep.append(ch)
        else:
            keep.append("_")
    return "".join(keep).strip() or "sheet"


def convert_xlsx_to_csv(input_path: str, sheet: str | None, out: str | None, out_dir: str | None) -> list[str]:
    wb = load_workbook(filename=input_path, data_only=True, read_only=True)

    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}. Available: {', '.join(wb.sheetnames)}")
        sheets = [sheet]
    else:
        sheets = wb.sheetnames

    outputs: list[str] = []
    for sheet_name in sheets:
        ws = wb[sheet_name]

        if out:
            if len(sheets) > 1:
                raise ValueError("--out can only be used when converting a single sheet")
            out_path = out
        else:
            base = os.path.splitext(os.path.basename(input_path))[0]
            fname = _safe_filename(f"{base} - {sheet_name}.csv")
            target_dir = out_dir or os.path.dirname(input_path) or "."
            out_path = os.path.join(target_dir, fname)

        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        outputs.append(out_path)

    return outputs


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert XLSX to CSV")
    parser.add_argument("input", help="Path to .xlsx file")
    parser.add_argument("--sheet", help="Sheet name to export (default: all sheets)")
    parser.add_argument("--out", help="Output CSV path (only for single sheet)")
    parser.add_argument("--out-dir", help="Output directory (default: same as input)")

    args = parser.parse_args()

    try:
        outputs = convert_xlsx_to_csv(args.input, args.sheet, args.out, args.out_dir)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    for p in outputs:
        print(p)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
