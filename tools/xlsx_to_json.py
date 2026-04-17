#!/usr/bin/env python3
import argparse
import json
import os
from typing import Any, Dict, List, Optional

import openpyxl


def normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return str(value).strip() or None


def make_unique_headers(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    unique: List[str] = []
    for h in headers:
        base = h
        if base not in seen:
            seen[base] = 1
            unique.append(base)
            continue
        seen[base] += 1
        unique.append(f"{base}_{seen[base]}")
    return unique


def find_header_row(ws, max_rows: int = 30) -> Optional[int]:
    best_row = None
    best_count = 0
    for r in range(1, max_rows + 1):
        count = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if normalize_header(v):
                count += 1
        if count > best_count:
            best_count = count
            best_row = r
    return best_row


def row_is_empty(values: List[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def build_records(ws, header_row: int, headers: List[str]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if row_is_empty(values):
            continue
        record = {headers[i]: values[i] for i in range(len(headers))}
        records.append(record)
    return records


def build_columns(ws, header_row: int, headers: List[str]) -> Dict[str, List[Any]]:
    columns: Dict[str, List[Any]] = {h: [] for h in headers}
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if row_is_empty(values):
            continue
        for i, h in enumerate(headers):
            columns[h].append(values[i])
    return columns


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert XLSX sheet to JSON")
    parser.add_argument(
        "-i",
        "--input",
        default="input/source/Master Data.xlsx",
        help="Input .xlsx file",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="outputs/Master_Data.json",
        help="Output .json file",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        default="Master_KB",
        help="Sheet name (default: Master_KB; falls back to active if not found)",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=0,
        help="Header row index (1-based). Use 0 to auto-detect.",
    )
    parser.add_argument(
        "--orient",
        choices=["records", "columns"],
        default="records",
        help="JSON orientation: records (rows) or columns",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active

    header_row = args.header_row if args.header_row > 0 else find_header_row(ws)
    if not header_row:
        raise ValueError("Header row not found. Provide --header-row.")

    headers_raw = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers_norm: List[str] = []
    for i, v in enumerate(headers_raw, start=1):
        h = normalize_header(v)
        headers_norm.append(h if h else f"column_{i}")
    headers = make_unique_headers(headers_norm)

    if args.orient == "records":
        data = build_records(ws, header_row, headers)
    else:
        data = build_columns(ws, header_row, headers)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
